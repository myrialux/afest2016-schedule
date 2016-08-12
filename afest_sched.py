#!/usr/bin/env python

"""
Code and classes to handle keeping the AnimeFest 2016 app's schedule in-sync with the version from the web site.
"""

import sys
import argparse
import csv
import re
import unicodedata
from datetime import datetime
from openpyxl import *


RETURN_VALUE_SUCCESS               = 0
RETURN_VALUE_INVALID_PARAMETER     = 1
RETURN_VALUE_WORKBOOK_ERROR        = 2
RETURN_VALUE_EXCESSIVE_ID_USE      = 3
RETURN_VALUE_SPLIT_EVENTS_SAME_DAY = 4

ATTENDIFY_SCHEDULE_SHEET_NAME = "Schedule"

ATTENDIFY_DESC_COL_INDEX = 4
ATTENDIFY_ID_COL_INDEX = 7

ATTENDIFY_DATE_FORMAT = "%m/%d/%Y"

DIFF_KEY_ADDED = "added"
DIFF_KEY_DELETED = "deleted"
DIFF_KEY_CHANGED = "changed"
DIFF_KEY_MATCHED = "matched"


# Common substitutions to cope with unicode
AFEST_DESC_SUBS = [(u"\u2019", "'"), (u"\u2014", "-"), (u"\u201c", "\""), (u"\u201d", "\""), (u"\u2013", "-"), (u"\u2011", "-"), (u"\u2122", "(TM)"), (u"\u0219", "s"), (u"\u2018", "'"), (u"\u2026", "..."), (u"\u00a0", " "), (u"\u000b", " ")]


class AFestEvent:
    """Model object for events, whether from the AFest schedule or Attendify."""

    def load_from_attendify(self, row):
        self.title = row[0].value.strip()
        self.date = row[1].value.strip()
        self.start_time = row[2].value.strip()
        self.end_time = row[3].value.strip()
        self.desc = row[ATTENDIFY_DESC_COL_INDEX].value.strip()
        self.desc = re.sub("&nbsp;", " ", self.desc)
        self.desc = re.sub("<br>", "", self.desc)
        self.location = row[5].value.strip()
        self.track = (row[6].value or "").strip()
        self.attendify_id = row[ATTENDIFY_ID_COL_INDEX].value.strip()

        regex = re.compile(r".*(\[afestid:(.+)\])$", re.IGNORECASE)
        match = regex.search(self.desc)
        if match:
            self.afest_id = match.group(2)

            # Trim the AFest ID so the Attendify event's desc matches
            full_id_str = match.group(1)
            trimmed_desc = self.desc[:-len(full_id_str)].strip()
            self.desc = trimmed_desc
        else:
            self.afest_id = None

    def load_from_afest(self, row):
        self.title = row["Session Title"].strip()
        self.date = row["Date"].strip()
        self.start_time = row["Start Time"].strip()
        self.end_time = row["End Time"].strip()
        if self.end_time == "00:00":
            # Treat ending at midnight as just before, to help match with Attendify
            self.end_time = "23:59"
        self.desc = unicode(row["Description"], "utf-8").strip()
        for sub in AFEST_DESC_SUBS:
            self.desc = self.desc.replace(sub[0], sub[1])
        self.location = row["Location"].strip()
        self.track = row["Track Title"].strip()
        self.attendify_id = row["UID"].strip()
        self.afest_id = row["id_schedule_block"].strip()

    def is_match(self, other):
        """Returns whether the other event matches this one for purposes of copying the AFest ID over.
        Not the same as an equality function, because we don't care about the track or description.
        """

        # Common checks first
        if (self.title != other.title) or (self.location != other.location):
            return False

        if (self.date == other.date):
            # If the days match then the start time needs to, and the end time either needs to match, or needs to be the end of the day. This helps with the first portion of events split over midnight.
            return (self.start_time == other.start_time) and ((self.end_time == other.end_time) or (self.end_time == "23:59"))
        else:
            # If this event is on the day after the other one, starts at midnight, and its end time matches, then this is the second part of a split-over-midnight event.
            thisDate = datetime.strptime(self.date, ATTENDIFY_DATE_FORMAT)
            otherDate = datetime.strptime(other.date, ATTENDIFY_DATE_FORMAT)
            return ((thisDate - otherDate).days == 1) and (self.start_time == "00:00") and (self.end_time == other.end_time)


def open_attendify_schedule(file_name):
    wb = load_workbook(file_name)

    schedule_sheet = wb[ATTENDIFY_SCHEDULE_SHEET_NAME]
    if not schedule_sheet:
        ArgumentParser.exit(RETURN_VALUE_WORKBOOK_ERROR, "ERROR - Schedule sheet not found in workbook")
    
    return wb


def iter_attendify_schedule_rows(sheet):
    """Returns an iterator for the schedule rows in the given sheet."""

    events_range = "A6:H" + str(sheet.max_row)
    return sheet.iter_rows(range_string=events_range)


def load_attendify_events(file_name):
    wb = open_attendify_schedule(file_name)
    schedule_sheet = wb[ATTENDIFY_SCHEDULE_SHEET_NAME]

    events = []

    for row in iter_attendify_schedule_rows(schedule_sheet):
        event = AFestEvent()
        event.load_from_attendify(row)
        events.append(event)

    return events


def load_afest_events(file_name):
    events = []

    with open(file_name) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            event = AFestEvent()
            event.load_from_afest(row)
            events.append(event)

    return events


def add_afest_id_to_attendify(workbook, attendify_id, afest_id):
    schedule_sheet = workbook[ATTENDIFY_SCHEDULE_SHEET_NAME]

    for row in iter_attendify_schedule_rows(schedule_sheet):
        if row[ATTENDIFY_ID_COL_INDEX].value.strip() == attendify_id:
            row[ATTENDIFY_DESC_COL_INDEX].value += "\n\n[afestid:{0}]".format(afest_id)
            break


def merge_events(event1, event2):
    if event1.date == event2.date:
        ArgumentParser.exit(RETURN_VALUE_SPLIT_EVENTS_SAME_DAY, "ERROR - Events to merge share the same date ({0}) - \"{1}\" and \"{2}\"".format(event1.date, event1.title, event2.title))

    date1 = datetime.strptime(event1.date, ATTENDIFY_DATE_FORMAT)
    date2 = datetime.strptime(event2.date, ATTENDIFY_DATE_FORMAT)

    if date1 < date2:
        startEvent = event1
        endEvent = event2
    else:
        startEvent = event2
        endEvent = event1

    result = AFestEvent()
    result.title = startEvent.title
    result.date = startEvent.date
    result.start_time = startEvent.start_time
    result.end_time = endEvent.end_time
    result.desc = startEvent.desc
    result.location = startEvent.location
    result.track = startEvent.track
    result.attendify_id = startEvent.attendify_id
    result.afest_id = startEvent.afest_id
    return result


def merge_split_events(events):
    """Takes the given list of events and returns a new list with all events that were split across midnight merged into a single event with the proper start and end time.
    The date field of merged events is the date of the start of the event.
    """

    by_afest_id = {}
    for event in events:
        id_list = []
        if by_afest_id.has_key(event.afest_id):
            id_list = by_afest_id[event.afest_id]
        id_list.append(event)
        by_afest_id[event.afest_id] = id_list

    result = []
    for afest_id in by_afest_id:
        id_list = by_afest_id[afest_id]
        if len(id_list) == 1:
            result.append(id_list[0])
        elif len(id_list) == 2:
            result.append(merge_events(id_list[0], id_list[1]))
        else:
            ArgumentParser.exit(RETURN_VALUE_EXCESSIVE_ID_USE, "ERROR - {0} event(s) for AFest ID {1}".format(len(id_list), afest_id))
    
    return result


def add_ids_to_attendify(args):
    afest_events = load_afest_events(args.afest_file)
    attendify_events = load_attendify_events(args.attendify_file)

    print("AFest: {0}  Attendify: {1}".format(len(afest_events), len(attendify_events)))

    workbook = open_attendify_schedule(args.attendify_file)

    exact_matches = 0
    title_matches = 0
    for at_event in attendify_events:
        if at_event.afest_id:
            continue

        # Exact matches first
        matched = False
        for af_event in afest_events:
            if at_event.is_match(af_event):
                add_afest_id_to_attendify(workbook, at_event.attendify_id, af_event.afest_id)

                matched = True
                exact_matches += 1
                break
        if matched:
            continue
        
        # If the event has exactly one match against the titles in the AFest schedule, use that one.
        matched_afest_ids = []
        for af_event in afest_events:
            if at_event.title == af_event.title:
                matched_afest_ids.append(af_event.afest_id)
        if len(matched_afest_ids) == 1:
            add_afest_id_to_attendify(workbook, at_event.attendify_id, matched_afest_ids[0])
            title_matches += 1

    workbook.save(args.attendify_file + ".new")
    
    print("Exact Matches: {0}  Title Matches: {1}".format(exact_matches, title_matches))


def check_afest_ids_in_attendify(args):
    attendify_events = load_attendify_events(args.attendify_file)

    missing_ids = 0
    for event in attendify_events:
        if not event.afest_id:
            print("{0}: {1} {2}-{3} {4}".format(event.title, event.date, event.start_time, event.end_time, event.location))
            missing_ids += 1

    print("Total events: {0}  Missing IDs: {1}".format(len(attendify_events), missing_ids))


def diff_event_lists(left, right):
    """Does the actual diff between the left (old, Attendify) and right (new, AFest) schedules. Returns a dictionary of event objects with keys to lists of added, deleted, changed, and matched items.
    In the case of changed items, each entry is a tuple where the first item is the old event and the second the new event.
    N.B. - The lists must be sorted by AFest ID before the call to this function.
    Went with indexes instead of Python iterators because the all the one-line try..catch blocks were getting ugly.
    """

    result = {DIFF_KEY_ADDED: [], DIFF_KEY_DELETED: [], DIFF_KEY_CHANGED: {}, DIFF_KEY_MATCHED: []}

    left_index = 0
    right_index = 0
    while (left_index < len(left)) and (right_index < len(right)):
        current_left = left[left_index]
        current_right = right[right_index]

        if (current_left.afest_id < current_right.afest_id):
            result[DIFF_KEY_DELETED].append(current_left)

            left_index += 1
        elif (current_left.afest_id > current_right.afest_id):
            result[DIFF_KEY_ADDED].append(current_right)

            right_index += 1
        else:
            changes = {}
            if current_left.date != current_right.date:
                changes["date"] = (current_left.date, current_right.date)
            if current_left.start_time != current_right.start_time:
                changes["start_time"] = (current_left.start_time, current_right.start_time)
            if current_left.end_time != current_right.end_time:
                changes["end_time"] = (current_left.end_time, current_right.end_time)
            if current_left.title != current_right.title:
                changes["title"] = (current_left.title, current_right.title)
            if current_left.location != current_right.location:
                changes["location"] = (current_left.location, current_right.location)
            norm_left_desc = unicodedata.normalize("NFC", current_left.desc)
            norm_right_desc = unicodedata.normalize("NFC", current_right.desc)
            if norm_left_desc != norm_right_desc:
                changes["desc"] = (norm_left_desc, norm_right_desc)
            if current_left.track != current_right.track:
                changes["track"] = (current_left.track, current_right.track)

            if len(changes) > 0:
                result[DIFF_KEY_CHANGED][current_left.afest_id] = changes
            else:
                result[DIFF_KEY_MATCHED].append(current_left)

            left_index += 1
            right_index += 1

    # Remaining items in the left list have been deleted
    while (left_index < len(left)):
        current_left = left[left_index]
        result[DIFF_KEY_DELETED].append(current_left)
        left_index += 1
    
    # Remaining items in the right list are new
    while (right_index < len(right)):
        current_right = right[right_index]
        result[DIFF_KEY_ADDED].append(current_right)
        right_index += 1

    return result


def diff_schedules(args):
    if len(args.attendify_files) == 0:
        ArgumentParser.exit(RETURN_VALUE_INVALID_PARAMETER, "ERROR - No Attendify schedules provided")

    afest_events = load_afest_events(args.afest_file)
    afest_events.sort(cmp=lambda l,r: cmp(l.afest_id, r.afest_id))

    # Cat all the Attendify events together
    attendify_events = []
    for attendify_file in args.attendify_files:
        attendify_events.extend(load_attendify_events(attendify_file))
    original_attendify_count = len(attendify_events)
    attendify_events = merge_split_events(attendify_events)
    attendify_events.sort(cmp=lambda l,r: cmp(l.afest_id, r.afest_id))

    print("AFest Events: {0}  Attendify: {1} ({2} pre-merge)".format(len(afest_events), len(attendify_events), original_attendify_count))

    deltas = diff_event_lists(attendify_events, afest_events)
    print("Added: {0}  Deleted: {1}  Changed: {2}  Matched: {3}".format(len(deltas[DIFF_KEY_ADDED]), len(deltas[DIFF_KEY_DELETED]), len(deltas[DIFF_KEY_CHANGED]), len(deltas[DIFF_KEY_MATCHED])))
    print("")

    if (len(deltas[DIFF_KEY_ADDED]) > 0):
        print("Added")
        print("-----")

        for a in deltas[DIFF_KEY_ADDED]:
            print("{0}, {1}, {2}, {3}-{4}, {5}, {6}: {7}".format(a.afest_id, a.title, a.date, a.start_time, a.end_time, a.location, a.track, a.desc))

        print("")

    if (len(deltas[DIFF_KEY_DELETED]) > 0):
        print("Deleted")
        print("-------")

        for d in deltas[DIFF_KEY_DELETED]:
            print("{0}, {1}, {2}, {3}-{4}, {5}".format(d.afest_id, d.title, d.date, d.start_time, d.end_time, d.location))

        print("")

    if (len(deltas[DIFF_KEY_CHANGED]) > 0):
        print("Changed")
        print("-------")

        for c in deltas[DIFF_KEY_CHANGED]:
            afest_event = filter(lambda x: x.afest_id == c, afest_events)[0]
            print("{0}, {1}, {2} {3}-{4}, {5} : {6}".format(c, afest_event.title, afest_event.date, afest_event.start_time, afest_event.end_time, afest_event.location, deltas[DIFF_KEY_CHANGED][c]))
        print("")


def main():
    parser = argparse.ArgumentParser(description="Show schedule data")
    subparsers = parser.add_subparsers()

    add_ids_parser = subparsers.add_parser("add_ids", help="Attempt to match AFest and Attendify events. Copy AFest IDs to matched Attendify events")
    add_ids_parser.add_argument("afest_file", help="AFest .csv schedule")
    add_ids_parser.add_argument("attendify_file", help="Attendify .xlsx schedule")
    add_ids_parser.set_defaults(func=add_ids_to_attendify)

    check_ids_parser = subparsers.add_parser("check_ids", help="Check the given Attendify schedule for missing AFest IDs.")
    check_ids_parser.add_argument("attendify_file", help="Attendify .xlsx schedule")
    check_ids_parser.set_defaults(func=check_afest_ids_in_attendify)

    diff_parser = subparsers.add_parser("diff", help="Diff the AFest schedule against the schedule(s) from Attendify")
    diff_parser.add_argument("afest_file", help="AFest .csv schedule")
    diff_parser.add_argument("attendify_files", nargs=argparse.REMAINDER, help="One or more Attendify .xlsx schedule files")
    diff_parser.set_defaults(func=diff_schedules)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
