#!/usr/bin/env python


import sys
import csv
import re
from openpyxl import *


RETURN_VALUE_SUCCESS           = 0
RETURN_VALUE_INVALID_PARAMETER = 1
RETURN_VALUE_WORKBOOK_ERROR    = 2

ATTENDIFY_SCHEDULE_SHEET_NAME = "Schedule"


class AFestEvent:
    """Model object for events, whether from the AFest schedule or Attendify."""

    title = ""
    date = ""
    start_time = ""
    end_time = ""
    desc = ""
    location = ""
    track = ""
    attendify_id = None
    afest_id = None

    def load_from_attendify(self, row):
        self.title = row[0].value.strip()
        self.date = row[1].value.strip()
        self.start_time = row[2].value.strip()
        self.end_time = row[3].value.strip()
        self.desc = row[4].value.strip()
        self.location = row[5].value.strip()
        self.track = row[6].value.strip()
        self.attendify_id = row[7].value.strip()

        regex = re.compile(r".*{afestid:(.+)}$", re.IGNORECASE)
        match = regex.search(self.desc)
        if match:
            self.afest_id = match.group(1)

    def load_from_afest(self, row):
        self.title = row["Session Title"]
        self.date = row["Date"]
        self.start_time = row["Start Time"]
        self.end_time = row["End Time"]
        self.desc = row["Description"]
        self.location = row["Location"]
        self.track = row["Track Title"]
        self.attendify_id = row["UID"]
        self.afest_id = row["id_schedule_block"]

    def is_match(self, other):
        """Returns whether the other event matches this one for purposes of copying the AFest ID over.
        Not the same as an equality function, because we don't care about the track or description.
        """

        return (self.date == other.date) and (self.start_time == other.start_time) and (self.end_time == other.end_time) and (self.location == other.location) and (self.title == other.title)


def open_attendify_schedule(file_name):
    wb = load_workbook(file_name)

    schedule_sheet = wb[ATTENDIFY_SCHEDULE_SHEET_NAME]
    if not schedule_sheet:
        print("ERROR - Schedule sheet not found in workbook")
        sys.exit(RETURN_VALUE_WORKBOOK_ERROR)
    
    return wb


def load_attendify_events(file_name):
    wb = open_attendify_schedule(file_name)
    schedule_sheet = wb[ATTENDIFY_SCHEDULE_SHEET_NAME]

    events = []

    events_range = "A6:H" + str(schedule_sheet.max_row)
    for row in schedule_sheet.iter_rows(range_string=events_range):
        event = AFestEvent()
        event.load_from_attendify(row)
        events.append(event)

    return events


def load_afest_events(file_name):
    events = []

    with open(file_name) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            event = AFestEvent()
            event.load_from_afest(row)
            events.append(event)

    return events


def add_ids_to_attendify(args):
    afest_events = load_afest_events(args.afest_file)
    attendify_events = load_attendify_events(args.attendify_file)

    print("AFest: {0}  Attendify: {1}".format(len(afest_events), len(attendify_events)))

    matches = 0
    title_matches = 0
    for at_event in attendify_events:
        matched = False
        for af_event in afest_events:
            if at_event.is_match(af_event):
                matched = True
                matches += 1
                break
        if matched:
            continue
        
        for af_event in afest_events:
            if at_event.title == af_event.title:
                title_matches += 1
                break 
    
    print("Matches: {0}  Title Matches: {1}".format(matches, title_matches))


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Show schedule data")
    subparsers = parser.add_subparsers()
    add_ids_parser = subparsers.add_parser("add_ids", help="Attempt to match AFest and Attendify events. Copy AFest IDs to matched Attendify events")
    add_ids_parser.add_argument("afest_file", help="AFest .csv schedule")
    add_ids_parser.add_argument("attendify_file", help="Attendify .xlsx schedule")
    add_ids_parser.set_defaults(func=add_ids_to_attendify)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
